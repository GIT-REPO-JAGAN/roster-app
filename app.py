import os
import json
import uuid
import re
import pandas as pd
from flask import Flask, request, jsonify, render_template, send_file
from groq import Groq
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from json_repair import repair_json

app = Flask(__name__, 
            static_folder='static', 
            static_url_path='/static', 
            template_folder='templates')

app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# Absolute Paths for Codespaces/Cloud compatibility
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
OUTPUT_FOLDER = os.path.join(BASE_DIR, 'outputs')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def read_roster(filepath):
    """Safely reads roster and handles missing columns to prevent KeyError."""
    df = pd.read_excel(filepath, header=None)
    employees = []
    for i, row in df.iterrows():
        val = str(row[0]).strip() if pd.notna(row[0]) else ""
        if val in ("", "Month", "Date", "Day") or pd.isna(row[0]):
            continue
        if len(row) >= 2 and pd.notna(row[0]) and pd.notna(row[1]):
            employees.append({
                "name":     str(row[0]).strip(),
                "email":    str(row[1]).strip(),
                "skill":    str(row[2]).strip() if (len(row) > 2 and pd.notna(row[2])) else "General",
                "location": str(row[3]).strip() if (len(row) > 3 and pd.notna(row[3])) else "Remote",
            })
    return employees

def build_groq_prompt(employees, start_date, end_date, custom_prompt):
    """Universal prompt using the Compressed Array Format to prevent truncation."""
    emp_list = "\n".join([f"  - {e['name']} | {e['skill']} | {e['location']}" for e in employees])
    return f"""You are a Universal Workforce Scheduling Engine.
    TASK: Generate a complete shift schedule.
    DATE RANGE: {start_date} to {end_date}
    EMPLOYEES:
    {emp_list}

    ### USER DEFINED RULES & SHIFT LEGEND:
    {custom_prompt if custom_prompt else "Standard business hours, Mon-Fri, WO for weekends."}

    ### STRICT EXECUTION RULES:
    1. COMPRESSED FORMAT: Return the shifts as a simple array of strings for each employee.
    2. ZERO TRUNCATION: Every employee MUST have a shift assigned for EVERY date in the range.
    3. Return ONLY raw JSON. No markdown, no conversation.
    4. Every array must have exactly {(datetime.strptime(end_date, "%Y-%m-%d") - datetime.strptime(start_date, "%Y-%m-%d")).days + 1} entries.

    OUTPUT FORMAT:
    {{
      "schedule": {{
        "Employee Name": ["SHIFT1", "SHIFT2", "SHIFT3", ...]
      }}
    }}"""

def call_groq(api_key, prompt):
    client = Groq(api_key=api_key)
    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile", 
        messages=[{"role": "user", "content": prompt}],
        temperature=0.0, 
        max_tokens=8000,
    )
    return response.choices[0].message.content.strip()

def generate_excel(employees, schedule_data, start_date_str, end_date_str, output_path):
    start = datetime.strptime(start_date_str, "%Y-%m-%d")
    end = datetime.strptime(end_date_str, "%Y-%m-%d")
    dates = [start + timedelta(days=i) for i in range((end - start).days + 1)]

    wb = Workbook()
    ws = wb.active
    ws.title = "Shift Schedule"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    month_fills = [PatternFill("solid", fgColor="2E75B6"), PatternFill("solid", fgColor="70AD47")]
    white_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
                         top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"))

    for row in range(1, 4):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.cell(row=1, column=1).value = "Month"
    ws.cell(row=2, column=1).value = "Date"
    ws.cell(row=3, column=1).value = "Day"
    for col, lbl in enumerate(["Name", "Email", "Skill", "Location"], 1):
        ws.cell(row=3, column=col).value = lbl

    col_offset = 5
    for i, d in enumerate(dates):
        col = col_offset + i
        is_weekend = d.weekday() >= 5
        c2 = ws.cell(row=2, column=col, value=d.day)
        c2.fill = PatternFill("solid", fgColor="D9D9D9" if is_weekend else "2E75B6")
        c2.font = white_font if not is_weekend else Font(color="333333")
        c3 = ws.cell(row=3, column=col, value=d.strftime("%a"))
        c3.fill = PatternFill("solid", fgColor="D9D9D9" if is_weekend else "1F4E79")
        c3.font = white_font if not is_weekend else Font(color="555555")
        ws.column_dimensions[get_column_letter(col)].width = 5

    month_groups = {}
    for i, d in enumerate(dates):
        key = d.strftime("%B %Y")
        if key not in month_groups: month_groups[key] = {"start": i, "end": i}
        else: month_groups[key]["end"] = i

    month_fill_idx = 0
    for month_key, span in month_groups.items():
        start_col = col_offset + span["start"]
        end_col = col_offset + span["end"]
        if end_col > start_col: ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        mc = ws.cell(row=1, column=start_col, value=month_key)
        mc.font = white_font
        mc.fill = month_fills[month_fill_idx % len(month_fills)]
        mc.alignment = Alignment(horizontal="center", vertical="center")
        month_fill_idx += 1

    # UPDATED COLORS: Added E1 and E2 to match your shift legend
    generic_colors = {
        "WO": "F2F2F2", "PL": "FCE5CD", "SL": "EA9999", "H": "FFE599",
        "E1": "D9D2E9", "E2": "B4A7D6", "G": "E2EFDA", "M": "DDEBF7", "A": "FFF2CC", "N": "F4CCCC"
    }

    for row_idx, emp in enumerate(employees):
        row = 4 + row_idx
        for col, val in enumerate([emp["name"], emp["email"], emp["skill"], emp["location"]], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = thin_border
        
        shifts = schedule_data.get(emp["name"], [])
        for i, d in enumerate(dates):
            col = col_offset + i
            if i < len(shifts):
                shift_code = shifts[i]
            else:
                shift_code = "WO" if d.weekday() >= 5 else "G"
            
            c = ws.cell(row=row, column=col, value=shift_code)
            c.fill = PatternFill("solid", fgColor=generic_colors.get(shift_code, "FFFFFF"))
            c.border = thin_border

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.freeze_panes = "E4"
    wb.save(output_path)

@app.route("/")
def index(): return render_template("index.html")

@app.route("/api/generate", methods=["POST"])
def generate():
    api_key = request.form.get("api_key", "").strip()
    start_date = request.form.get("start_date", "").strip()
    end_date = request.form.get("end_date", "").strip()
    custom_prompt = request.form.get("custom_prompt", "").strip()
    file = request.files.get("roster_file")

    if not all([api_key, start_date, end_date, file]):
        return jsonify({"error": "All fields are required."}), 400

    filename = secure_filename(file.filename)
    upload_path = os.path.join(UPLOAD_FOLDER, f"{uuid.uuid4()}_{filename}")
    file.save(upload_path)

    try:
        employees = read_//Crosture employees = read_roster(upload_path)
        if not employees: return jsonify({"error": "No valid employee data found."}), 400
        prompt = build_groq_prompt(employees, start_date, end_date, custom_prompt)
        groq_response = call_groq(api_key, prompt)
        raw = groq_response.strip()
        if raw.startswith("```"): raw = re.sub(r'^```json\s*|```$', '', raw, flags=re.MULTILINE)
        repaired_json_str = repair_json(raw) 
        schedule_data = json.loads(repaired_json_str).get("schedule", {})
        output_id = str(uuid.uuid4())
        output_path = os.path.join(OUTPUT_FOLDER, f"schedule_{output_id}.xlsx")
        generate_excel(employees, schedule_data, start_date, end_date, output_path)
        return jsonify({"success": True, "download_id": output_id, "employee_count": len(employees)})
    except Exception as e:
        return jsonify({"error": f"Processing Error: {str(e)}"}), 500
    finally:
        if os.path.exists(upload_path): os.remove(upload_path)

@app.route("/api/download/<download_id>")
def download(download_id):
    safe_id = secure_filename(download_id)
    path = os.path.join(OUTPUT_FOLDER, f"schedule_{safe_id}.xlsx")
    if os.path.exists(path):
        return send_file(path, as_attachment=True, download_name="Roster_Output.xlsx")
    return jsonify({"error": "File not found"}), 404

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
